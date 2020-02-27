#!/usr/bin/python3
'''
**********************************************************************************
Description:  Extract occupancy stats from plaidbench runs.
Author:       Vang Thao
Created:      December 30, 2019
Last Update:  December 30, 2019
**********************************************************************************

OUTPUT:
    This script takes in data from plaidbench runs and output a spreadsheet
    containing the average occupancy for each benchmark and the overall
    average occupancy.
        Spreadsheet 1: occupancy.xlsx

Requirements:
    - python3
    - pip3
    - openpyxl (sreadsheet module, installed using pip3)
    - patch to print out occupancy

HOW TO USE:
    1.) Run a plaidbench benchmarks with run-plaidbench.sh to generate a
        directory containing the results for the run.
    2.) Move the directory into a separate folder containing only the
        directories generated by the script.
    3.) Pass the path to the folder as an input to this script with
        the -i option.

Example:
    ./get-occupancy.py -i /home/tom/plaidbench-runs
    
    where plaidbench-runs/ contains
        plaidbench-optsched-01/
        plaidbench-optsched-02/
        ...
        plaidbench-amd-01/
        ...
'''

import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse

RE_OCCUPANCY = re.compile('Final occupancy for function (.*):(\d+)')

# Contains all of the stats
benchStats = {}

# List of benchmark names
benchmarks = [
    'densenet121',
    'densenet169',
    'densenet201',
    'inception_resnet_v2',
    'inception_v3',
    'mobilenet',
    'nasnet_large',
    'nasnet_mobile',
    'resnet50',
    'vgg16',
    'vgg19',
    'xception',
    'imdb_lstm',
]

# Ignore these functions
# They are outputted before scheduling
ignore = [
    'copyBufferRect',
    'copyBufferRectAligned',
    'copyBuffer',
    'copyBufferAligned',
    'fillBuffer',
    'copyBufferToImage',
    'copyImageToBuffer',
    'copyImage',
    'copyImage1DA',
    'fillImage',
    'scheduler'
]

def parseStats(inputFolder, ignoreFolders):
    scanDirPath = os.path.abspath(inputFolder)
    
    # Get name of all directories in the specified folder
    subfolders = [f.name for f in os.scandir(path=scanDirPath) if f.is_dir() ]

    # For each folder
    for folderName in subfolders:
        if folderName in ignoreFolders:
            continue
        name = folderName.split('-')

        # Get the run number from the end
        # of the folder name
        runNumber = name[-1]

        # Get the name of the run
        # and exclude the run number
        nameOfRun = '-'.join(name[:-1])
            
        # Create an entry in the stats for the
        # name of the run
        if (nameOfRun not in benchStats):
            benchStats[nameOfRun] = {}

        for bench in benchmarks:
            # Get the path to the log file
            currentPath = os.path.join(inputFolder, folderName)
            currentPath = os.path.join(currentPath, bench)
            currentLogFile = os.path.join(currentPath, bench + '.log')
            stats = {}
            stats['average'] = 0.0
            stats['total'] = 0.0
            stats['numKernel'] = 0

            # First check if log file exists.
            if (os.path.exists(currentLogFile)):
                # Open log file if it exists.
                with open(currentLogFile) as file:
                    for line in file:
                        # Match the line that contain occupancy stats
                        getOccupancyStats = RE_OCCUPANCY.match(line)
                        if (getOccupancyStats):
                            # Get the kernel name
                            kernelName = getOccupancyStats.group(1)

                            # Ignore these function
                            if (kernelName in ignore):
                                continue

                            # Get occupancy
                            occupancy = int(getOccupancyStats.group(2))
                            
                            # Used for averaging
                            stats['total'] += occupancy
                            stats['numKernel'] += 1
            else:
                print('Cannot find log file for {} run {} benchmark {}.'.format(nameOfRun, runNumber, bench))
            
            if stats['numKernel'] != 0:
                stats['average'] = stats['total'] / stats['numKernel']

            # Save stats
            benchStats[nameOfRun][bench] = stats

def printStats():
    for nameOfRun in benchStats:
        print('{}'.format(nameOfRun))
        total = 0.0
        kernel = 0
        for bench in benchStats[nameOfRun]:
            print('    {} : {:.2f}'.format(bench, benchStats[nameOfRun][bench]['average']))
            total += benchStats[nameOfRun][bench]['total']
            kernel += benchStats[nameOfRun][bench]['numKernel']
        if kernel != 0:
            print('  Average: {:.2f}'.format(total/kernel))

def createSpreadsheets(output):
    if 'xls' not in output[-4:]:
        output += '.xlsx'
    
    # Create new excel worksheet
    wb = Workbook()

    # Grab the active worksheet
    ws = wb.active

    # Insert title and benchmark names
    ws['A1'] = 'Benchmarks'
    ws['A1'].font = Font(bold=True)

    row = 3
    for bench in benchmarks:
        ws['A' + str(row)] = bench
        row += 1
        
    ws['A' + str(row)] = 'Average'
    ws['A' + str(row)].font = Font(bold=True)

    # Stats entry
    col = 'B'
    for nameOfRun in benchStats:
        row = 1
        ws[col + str(row)] = nameOfRun
        row = 2
        ws[col+str(row)] = 'Occupancy'
        row = 3
        
        total = 0.0
        kernel = 0
        for bench in benchmarks:
            ws[col+str(row)] = benchStats[nameOfRun][bench]['average']
            total += benchStats[nameOfRun][bench]['total']
            kernel += benchStats[nameOfRun][bench]['numKernel']
            row += 1
        ws[col+str(row)] = total/kernel
        
        # Convert column char to ASCII value
        # then increment it and convert
        # back into char. Used to go to next
        # column for next test run.
        col = chr(ord(col)+1)

    wb.save(output)

def main(args):
    # Parse folders to ignore into a list
    ignoreFolders = args.ignoreFolders.split(',')

    # Start stats collection
    parseStats(args.inputFolder, ignoreFolders)

    if args.verbose:
        printStats()

    if not args.disable:
        createSpreadsheets(args.output)    

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Script to extract occupancy data. \
                                     Requires patch to print occupancy.', \
                                      formatter_class=argparse.ArgumentDefaultsHelpFormatter)

    parser.add_argument('--verbose', '-v',
                        action='store_true', default=False,
                        dest='verbose',
                        help='Print average occupancy to terminal')

    parser.add_argument('--output', '-o',
                        default='occupancy',
                        dest='output',
                        help='Output spreadsheet filepath')

    parser.add_argument('--disable', '-d',
                        action='store_true', default=False,
                        dest='disable',
                        help='Disable spreadsheet output.')

    parser.add_argument('--input', '-i',
                        default='.',
                        dest='inputFolder',
                        help='The path to scan for benchmark directories')

    parser.add_argument('--ignore',
                        type=str,
                        default='',
                        dest='ignoreFolders',
                        help='List of folders to ignore separated by semi-colon')

    args = parser.parse_args()

    main(args)